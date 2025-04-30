from odoo import models, fields, api, _  # type: ignore
from odoo.addons.is_dynacase2odoo.models.is_param_project import GESTION_J, TYPE_DOCUMENT, MODELE_TO_TYPE, TYPE_TO_FIELD  # type: ignore
from odoo.exceptions import AccessError, ValidationError, UserError  # type: ignore
from datetime import datetime, timedelta, date
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Font, Color, Fill, Alignment,PatternFill
from copy import copy
import pytz
import calendar
from random import *
import base64
import cairo
from PIL import Image
from io import BytesIO
import time
import os
from matplotlib.colors import to_rgb
import codecs
import xml.etree.ElementTree as ET


_CSS_COLOR={
    'lavender'      : "#e6e6fa",
    'cornflowerblue': "#6495ed",
    'red'           : "#ff0000",
    'gray'          : "#000000",
    'springgreen'   : "#00ff7f",
    'orange'        : "#ffa500",
}


class IsGanttPdfExcel(models.Model):
    _name        = "is.gantt.pdf.excel"
    _description = "Modèle Excel pour le Gantt"

    name       = fields.Char("Nom du modèle", required=True)
    modele_ids = fields.Many2many('ir.attachment', string='Modèle Excel', required=True)


class IsGanttPdfSection(models.Model):
    _name        = "is.gantt.pdf.section"
    _description = "Sections du Gantt PDF/Excel"

    gantt_pdf_id = fields.Many2one('is.gantt.pdf', 'Gantt PDF', required=True, ondelete='cascade')
    section_id   = fields.Many2one("is.section.gantt", string="Section", required=True)
    afficher     = fields.Boolean("Afficher", default=True)


    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
           if 'gantt_pdf_id' not in vals:
                raise ValidationError("Il faut enregistrer ce document avant de modifier les sections")
        return super().create(vals_list)


class IsGanttPdf(models.Model):
    _name        = "is.gantt.pdf"
    _inherit=['mail.thread']
    _description = "Gantt PDF/Excel"
    _order = "name"


    name                      = fields.Char("Document", compute='_compute_name',store=True, readonly=True)
    type_document             = fields.Selection(TYPE_DOCUMENT,string="Type de document", default="Moule", required=True, tracking=True)
    modele_excel_id           = fields.Many2one("is.gantt.pdf.excel"       , string="Modèle Excel", tracking=True)
    moule_id                  = fields.Many2one("is.mold"                  , string="Moule", tracking=True)
    dossierf_id               = fields.Many2one("is.dossierf"              , string="Dossier F", tracking=True)
    dossier_modif_variante_id = fields.Many2one("is.dossier.modif.variante", string="Dossier Modif / Variante", tracking=True)
    dossier_article_id        = fields.Many2one("is.dossier.article"       , string="Dossier article", tracking=True)
    dossier_appel_offre_id    = fields.Many2one("is.dossier.appel.offre"   , string="Dossier appel d'offre", tracking=True)
    date_debut                = fields.Date("Date début", tracking=True)
    date_fin                  = fields.Date("Date fin", tracking=True)
    bordure_jour              = fields.Boolean("Bordure jour", default=False, tracking=True)
    logo_droite               = fields.Image(string="Logo de droite", tracking=True)
    format_fichier            = fields.Selection([
        ("png"  , "PNG"),
        ("pdf"  , "PDF"),
        ("svg"  , "SVG"),
        ("xlsx" , "Excel"),
    ], string="Format fichier", default="png", tracking=True)
    section_ids = fields.One2many('is.gantt.pdf.section', 'gantt_pdf_id')


    @api.onchange('type_document','moule_id','dossierf_id','dossier_modif_variante_id','dossier_article_id','dossier_appel_offre_id')
    def onchange_moule(self):
        for obj in self:
            items,titre,jour_fermeture_ids,markers = obj.get_taches()
            lines=[]
            ids=[]
            for item in items:
                model  = item.get('model')
                res_id = int(item.get('res_id'))
                if model=='is.section.gantt' and res_id>0:
                    if res_id not in ids:
                        ids.append(res_id)
            for id in ids:
                section = self.env['is.section.gantt'].browse(id)
                if section:
                    vals={
                        'section_id': id,
                        'afficher'  : section.gantt_pdf,
                    }
                    lines.append([0,0,vals])
            obj.section_ids = False
            obj.section_ids = lines


    @api.depends('type_document', 'moule_id', 'dossierf_id', 'dossier_modif_variante_id', 'dossier_article_id', 'dossier_appel_offre_id')
    def _compute_name(self):
        for obj in self:
            name=""
            if obj.type_document=="Moule":
                name = obj.moule_id.name
            if obj.type_document=="Dossier F":
                name = obj.dossierf_id.name
            if obj.type_document=="Article":
                name = obj.dossier_article_id.code_pg
            if obj.type_document=="Dossier Modif Variante":
                name = obj.dossier_modif_variante_id.demao_num
            if obj.type_document=="dossier_appel_offre":
                name = obj.dossier_appel_offre_id.dao_num
            obj.name = name


    def get_gantt_pdf_id(self,dossier_model,dossier_id):
        type_document = dict(MODELE_TO_TYPE).get(dossier_model)
        field_id      = dict(TYPE_TO_FIELD).get(type_document)
        if field_id=='idmoule':
            field_id='moule_id'
        domain=[
            ('create_uid'   ,'=', self.env.user.id),
            ('type_document','=', type_document),
            (field_id       ,'=', dossier_id),
        ]
        gantt_pdf_id=False
        for obj in self:
            lines=self.env['is.gantt.pdf'].search(domain, limit=1) #, order="date_fin_gantt"
            for line in lines:
                gantt_pdf_id = line.id
            if not gantt_pdf_id:
                vals={
                    'type_document': type_document,
                    field_id       : dossier_id,
                }
                gantt_pdf = self.env['is.gantt.pdf'].create(vals)
                gantt_pdf.onchange_moule()
                gantt_pdf_id = gantt_pdf.id
        return gantt_pdf_id


    def get_taches(self, section_ids=False,gantt_pdf=False):
        "Recherche des tâches en fonction des paramètres"
        for obj in self:
            items=[]
            domain=[]
            jour_fermeture_ids=[]
            markers=[]
            if gantt_pdf:
                domain.append(('gantt_pdf','=',True))
            if section_ids:
                domain.append(('section_id','in',section_ids))
            titre="?"
            if obj.type_document=="Moule" and obj.moule_id:
                domain.append(('idmoule','=',obj.moule_id.id))
                titre="%s - %s"%(obj.moule_id.name,obj.moule_id.designation)
            if obj.type_document=="Dossier F" and obj.dossierf_id:
                domain.append(('dossierf_id','=',obj.dossierf_id.id))
                titre="%s - %s"%(obj.dossierf_id.name,obj.dossierf_id.designation)
            if obj.type_document=="dossier_appel_offre" and obj.dossier_appel_offre_id:
                domain.append(('dossier_appel_offre_id','=',obj.dossier_appel_offre_id.id))
                titre=obj.dossier_appel_offre_id.dao_num
            if obj.type_document=="Dossier Modif Variante" and obj.dossier_modif_variante_id:
                domain.append(('dossier_modif_variante_id','=',obj.dossier_modif_variante_id.id))
                titre=obj.dossier_modif_variante_id.demao_num
            now = datetime.now(pytz.timezone('Europe/Paris')).strftime('%d/%m/%y %H:%M')
            titre="%s du %s"%(titre,now)
            if domain!=[]:
                res=self.env['is.doc.moule'].get_dhtmlx(domain=domain)
                items              = res['items']
                jour_fermeture_ids = res['jour_fermeture_ids']
                markers            = res['markers']
            return items,titre,jour_fermeture_ids,markers


    def get_sections(self):
        ids=[]
        for obj in self:
            for line in obj.section_ids:
                if line.afficher:
                    ids.append(line.section_id.id)
        return ids


    def generer_pdf_action(self):
        for obj in self:
            section_ids=obj.get_sections()
            items,titre,jour_fermeture_ids,markers = obj.get_taches(section_ids=section_ids, gantt_pdf=True)

            #** Calcul start_date *********************************************
            for item in items:
                end_date = item.get('end_date')
                duration =  item.get('duration')
                if end_date and  duration:
                    end_date   = datetime.strptime(end_date, '%Y-%m-%d 00:00:00')
                    start_date = end_date - timedelta(days=duration)
                    item['start_date'] = start_date
                    item['end_date']   = end_date
            #******************************************************************

            #** Recherche des enfants de chaque niveau ************************
            def get_enfants(items,parent,niveau):
                "Recherche des enfants du parent indiqué"
                for item in items:
                    if item.get('parent')==parent:
                        item['niveau'] = niveau
                        enfants.append(item)
                        get_enfants(items, item['id'], niveau+1)
                return
            enfants=[]
            get_enfants(items,False,0)
            #******************************************************************
 
            #** Calcul date de debut et de fin des parents ********************
            fin=False
            while fin==False:
                fin=True
                for enfant in enfants:
                    end_date = enfant.get('end_date')
                    if end_date:
                        for enfant2 in enfants:
                            if enfant2['id']==enfant.get('parent'):
                                if 'end_date' not in enfant2:
                                    enfant2['end_date']=end_date
                                    fin=False
                                if enfant2['end_date']<end_date:
                                    enfant2['end_date']=end_date
                                    fin=False
                    start_date = enfant.get('start_date')
                    if start_date:
                        for enfant2 in enfants:
                            if enfant2['id']==enfant.get('parent'):
                                if 'start_date' not in enfant2 or enfant2['start_date']==False:
                                    enfant2['start_date']=start_date
                                    fin=False
                                if enfant2['start_date']>start_date:
                                    enfant2['start_date']=start_date
                                    fin=False
            #******************************************************************

            #** Ajout des durations pour les parents **************************
            for enfant in enfants:
                duration=1
                if 'end_date' in enfant:
                    if 'duration' not in enfant or enfant['duration']==False:
                        duration = (enfant['end_date'] - enfant['start_date']).days
                        enfant['duration'] = duration
            #******************************************************************

            #** Ajout de la couleur de la section **************************
            for enfant in enfants:
                model  = enfant.get('model')
                res_id = int(enfant.get('res_id') or 0)
                color = "#1b81e8"
                o = self.env[model].browse(res_id)
                if o:
                    if model=='is.doc.moule' and o.section_id:
                        if o.section_id.color:
                            color = o.section_id.color
                    if model=='is.section.gantt':
                        if o.color:
                            color = o.color
                enfant['color'] = color
            #******************************************************************

            #** Recherche du nombre de jours **********************************
            def get_date(end_date,duration):
                start_date = False
                if end_date:
                    start_date = end_date - timedelta(days=duration)
                    #start_date = datetime.strptime(end_date, '%Y-%m-%d 00:00:00') - timedelta(days=duration)
                return start_date
            date_debut = date_fin = False
            items = enfants
            for item in items:
                duration   = item.get('duration')
                end_date   = get_date(item.get('end_date'),0)
                start_date = get_date(item.get('end_date'),duration)
                if start_date:
                    if not date_debut or date_debut>start_date:
                        date_debut =start_date
                if end_date:
                    if not date_fin or date_fin<end_date:
                        date_fin = end_date
            if obj.date_debut:
                date_debut=datetime.fromisoformat(obj.date_debut.isoformat()) 
            if obj.date_fin:
                date_fin=datetime.fromisoformat(obj.date_fin.isoformat()) 
            jour_nb=0
            if date_debut and date_fin:
                jour_nb = (date_fin-date_debut).days
            #******************************************************************

            #** Calcul du nombre de taches à afficher *************************
            nb_taches=0
            for item in items:
                duration   = item.get('duration')
                end_date   = get_date(item.get('end_date'),0)
                start_date = get_date(item.get('end_date'),duration)
                decal = 0
                if start_date:
                    decal = (start_date - date_debut).days
                if decal<0:
                    duration+=decal
                    decal=0
                if duration>0 and decal<jour_nb:
                    nb_taches+=1

            #** Paramètres du Gannt *******************************************
            file_extension = obj.format_fichier  # svg, png ou pdf
            file_name = obj.name
            file_name_with_extension = "%s.%s"%(file_name,file_extension)
            path = "/tmp/%s"%file_name_with_extension 
            tache_height        = 20
            jour_width          = 13
            grille_width        = 30*jour_width
            entete_height       = 4*tache_height # Entête pour placer le logo et un titre (ex : le moule)
            entete_table_height = 2*tache_height # Entête du tableau pour mettre les jours et les semaines

            WIDTH = grille_width + jour_nb*jour_width
            HEIGHT = entete_height + entete_table_height + nb_taches*tache_height
            #******************************************************************

            if file_extension!="xlsx":
                if file_extension=="svg":
                    surface = cairo.SVGSurface(path, WIDTH, HEIGHT)
                else:
                    surface = cairo.ImageSurface(cairo.FORMAT_RGB24,WIDTH,HEIGHT)
                ctx = cairo.Context(surface)

                def cairo_rectangle(ctx,x,y,width,height,line_width=0,line_rgb=(0,0,0),fill_rgb=False):
                    ctx.rectangle(x, y, width, height)
                    if fill_rgb:
                        ctx.set_source_rgb(fill_rgb[0],fill_rgb[1],fill_rgb[2])
                        ctx.fill()
                    if line_width>0:
                        ctx.rectangle(x, y, width, height)
                        ctx.set_source_rgb(line_rgb[0],line_rgb[1],line_rgb[2])
                        ctx.set_line_width(line_width) 
                        ctx.stroke() 
                    
                def cairo_show_text(ctx,x,y,font_rgb=(0,0,0),font_size=12,txt=""):
                    txt=txt or ''
                    ctx.set_source_rgb(font_rgb[0],font_rgb[1],font_rgb[2])
                    ctx.set_font_size(font_size)
                    ctx.select_font_face("Arial",cairo.FONT_SLANT_NORMAL,cairo.FONT_WEIGHT_NORMAL)
                    xbearing, ybearing, txt_width, txt_height, dx, dy = ctx.text_extents(txt)
                    ctx.move_to(x, y-txt_height/2)
                    ctx.show_text(txt)

                #** Fond gris clair pour toute la surface *************************
                cairo_rectangle(ctx,0,0,WIDTH,HEIGHT,fill_rgb=(0.95, 0.95, 0.95))

                #** Entete du Gantt ***********************************************
                cairo_rectangle(ctx,0,0,WIDTH,entete_height,fill_rgb=(1, 1, 1))
                cairo_show_text(ctx,220,entete_height-tache_height,font_size=32,txt=titre)

                #** Entete du tableau *********************************************
                cairo_rectangle(ctx,0,entete_height,WIDTH,tache_height,line_rgb=(0.8, 0.8, 0.8))
                #cairo_rectangle(ctx,0,entete_height+tache_height,WIDTH,tache_height,line_rgb=(0.8, 0.8, 0.8))

                #** Mois **********************************************************
                ladate = date_debut
                lesmois=[]
                for x in range(0,int(jour_nb)):
                    txt = ladate.strftime("%m/%Y")
                    if x==0 or ladate.day==1:
                        cemois=calendar.monthrange(ladate.year,ladate.month)
                        premier_du_mois=ladate.day
                        dernier_du_mois=cemois[1]
                        fin_mois = ladate + timedelta(days=dernier_du_mois)
                        delta = (fin_mois - date_fin).days
                        if delta>0:
                            dernier_du_mois = dernier_du_mois - delta
                        lesmois.append((txt,premier_du_mois,dernier_du_mois))
                    ladate += timedelta(days=1)
                decal=0
                for mois in lesmois:
                    nb_jours_mois = mois[2]-mois[1]+1 #dernier_du_mois - premier_du_mois
                    x      = grille_width+decal*jour_width
                    y      = entete_height
                    width  = nb_jours_mois *  jour_width
                    height = tache_height
                    cairo_rectangle(ctx,x,y,width,height,line_rgb=(0.8, 0.8, 0.8),line_width=0.2)
                    cairo_show_text(ctx,x+2,y+tache_height,txt=mois[0])
                    decal+=nb_jours_mois

                #** Semaines ******************************************************
                ladate = date_debut
                semaines=[]
                for x in range(0,int(jour_nb)):
                    #txt = ladate.strftime("S%W")
                    txt = ladate.strftime("S%V")
                    jour_semaine = ladate.isoweekday()
                    if x==0 or jour_semaine==1:
                        nb_jours_semaine = 7 - jour_semaine
                        debut_semaine = ladate.isoweekday()
                        fin_semaine   = 7
                        date_fin_semaine = ladate + timedelta(days=nb_jours_semaine)
                        delta = (date_fin_semaine - date_fin).days
                        if delta>0:
                            fin_semaine = fin_semaine - delta
                        semaines.append((txt,debut_semaine,fin_semaine))
                    ladate += timedelta(days=1)
                decal=0
                for semaine in semaines:
                    nb_jours_semaine = semaine[2]-semaine[1]+1
                    x      = grille_width+decal*jour_width
                    y      = entete_height+tache_height
                    width  = nb_jours_semaine *  jour_width
                    height = tache_height
                    cairo_rectangle(ctx,x,y,width,height,line_rgb=(0.8, 0.8, 0.8),line_width=0.2)
                    cairo_show_text(ctx,x+2,y+tache_height,txt=semaine[0])
                    decal+=nb_jours_semaine

                #** Ajout du tableau des tâches et de l'alternance des couleurs ***
                nb=0
                for item in items:
                    duration   = item.get('duration')
                    end_date   = get_date(item.get('end_date'),0)
                    start_date = get_date(item.get('end_date'),duration)
                    decal = 0
                    if start_date:
                        decal = (start_date - date_debut).days
                    if decal<0:
                        duration+=decal
                        decal=0
                    if duration>0 and decal<jour_nb:
                        x      = decal*jour_width+grille_width
                        y      = entete_height + entete_table_height + nb*tache_height
                        width  = duration*jour_width
                        height = tache_height
                        if nb%2:
                            fill_rgb=(227/255, 237/255, 252/255)
                            cairo_rectangle(ctx,0,y,WIDTH,tache_height,fill_rgb=fill_rgb,line_rgb=(0.8, 0.8, 0.8),line_width=0.2) # 1 ligne bleu toutes les 2 lignes
                        else:
                            cairo_rectangle(ctx,0,y,WIDTH,tache_height,line_rgb=(0.8, 0.8, 0.8),line_width=0.2)                   # 1 ligne blanche toutes les 2 lignes
                        nb+=1

                #** weekend en couleur ********************************************
                ladate = date_debut
                for ct in range(0,jour_nb):
                    if ladate.weekday() in (5,6):
                        #fill_rgb=to_rgb("#fadcb8")
                        fill_rgb=to_rgb("#d3d3d3")
                        x      = grille_width+ct*jour_width
                        y      = entete_height+tache_height*1
                        height = nb_taches*tache_height+tache_height
                        cairo_rectangle(ctx,x,y,jour_width,height,fill_rgb=fill_rgb)
                    ladate += timedelta(days=1)

                #** Ajout des vacances ********************************************
                ladate = date_debut
                for ct in range(0,jour_nb):
                    ladate_str = str(ladate)[0:10]
                    if ladate_str in jour_fermeture_ids:
                        couleur = jour_fermeture_ids[ladate_str][1]
                        fill_rgb=to_rgb("#faf5b8")
                        if couleur=='chine':
                            fill_rgb=to_rgb("#fadcb8")
                        x      = grille_width+ct*jour_width
                        y      = entete_height+tache_height*2
                        height = nb_taches*tache_height #+tache_height
                        cairo_rectangle(ctx,x,y,jour_width,height,fill_rgb=fill_rgb)
                    ladate += timedelta(days=1)

                #** Ligne verticale pour le now ***********************************
                ladate = date_debut
                now = date.today()
                for ct in range(0,jour_nb):
                    ladate_str = str(ladate)[0:10]
                    if ladate_str==str(now):
                        fill_rgb=to_rgb("#dc143c")
                        x      = grille_width+ct*jour_width
                        y      = entete_height #+tache_height*1
                        height = nb_taches*tache_height+tache_height*2
                        cairo_rectangle(ctx,x,y,jour_width/4,height,fill_rgb=fill_rgb)
                        cairo_show_text(ctx,x,y,txt='Now',font_rgb=fill_rgb)
                    ladate += timedelta(days=1)

                #** Lignes verticales pour les J **********************************
                ladate = date_debut
                for ct in range(0,jour_nb):
                    ladate_str = str(ladate)[0:10]
                    for marker in markers:
                        if marker['start_date'][0:10]==ladate_str:
                            fill_rgb=to_rgb("#dc143c")
                            x      = grille_width+ct*jour_width
                            y      = entete_height #+tache_height*1
                            height = nb_taches*tache_height+tache_height*2
                            cairo_rectangle(ctx,x,y,jour_width/4,height,fill_rgb=fill_rgb)
                            cairo_show_text(ctx,x,y,txt=marker['j'],font_rgb=fill_rgb)
                    ladate += timedelta(days=1)

                #** Ajout des tâches **********************************************
                nb=0
                for item in items:
                    duration   = item.get('duration')
                    end_date   = get_date(item.get('end_date'),0)
                    start_date = get_date(item.get('end_date'),duration)
                    decal = 0
                    if start_date:
                        decal = (start_date - date_debut).days
                    if decal<0:
                        duration+=decal
                        decal=0
                    if duration>0 and decal<jour_nb:
                        x      = decal*jour_width+grille_width
                        y      = entete_height + entete_table_height + nb*tache_height
                        width  = duration*jour_width

                        height = tache_height
                        txt = '%s %s'%('- '*item.get('niveau'),item.get('text'))
                        cairo_show_text(ctx,5,y+tache_height-1,txt=txt) # Nom de la tache à gauche
                        if obj.bordure_jour:
                            for ct in range(0,jour_nb):
                                cairo_rectangle(ctx,ct*jour_width+grille_width,y,jour_width,tache_height,line_rgb=(0.8, 0.8, 0.8),line_width=0.2) # Bordure des jours
                        else:
                            cairo_rectangle(ctx,0,y,grille_width,tache_height,line_rgb=(0.8, 0.8, 0.8),line_width=0.2) # Bordure du tableau des taches
                        color = item.get('color')
                        fill_rgb = to_rgb(color)
                        cairo_rectangle(ctx,x,y+3,width,height-6,fill_rgb=fill_rgb,line_rgb=(0, 0, 0),line_width=0.5) # Tache
                        fill_rgb=(0, 0, 0)
                        etat_class = item.get('etat_class')
                        if etat_class=='etat_a_faire':
                            fill_rgb=(1, 0, 0)
                        if etat_class=='etat_fait':
                            fill_rgb=(0, 1, 0)

                        # css_color={
                        #     'lavender'      : "#e6e6fa",
                        #     'cornflowerblue': "#6495ed",
                        #     'red'           : "#ff0000",
                        #     'gray'          : "#000000",
                        #     'springgreen'   : "#00ff7f",
                        #     'orange'        : "#ffa500",
                        # }
                        border_color = (item.get('border_color') or '').lower()
                        if border_color in _CSS_COLOR:
                            fill_rgb = to_rgb(_CSS_COLOR[border_color])

                        cairo_rectangle(ctx,x,y+3,tache_height/2,height-6,fill_rgb=fill_rgb)    # Rectangle Fait/Pas fait
                        cairo_show_text(ctx,x+tache_height,y+tache_height,txt=item.get('text')) # Nom de la tache sur le rectangle de la tache
                        nb+=1

                #** Logo au format PIL et redimmensionnement **********************
                logo_path="/tmp/logo-gantt-pdf.png"
                company = self.env.user.company_id
                f = open(logo_path,'wb')
                f.write(base64.b64decode(company.logo))
                f.close()
                image_logo = Image.open(logo_path)  
                width, height = image_logo.size 
                max_height = entete_height
                ratio = height/max_height
                new_width = int(width/ratio)
                image_logo_resize = image_logo.resize((new_width, max_height))

                if obj.logo_droite:
                    logo_path="/tmp/logo-droite.png"
                    f = open(logo_path,'wb')
                    f.write(base64.b64decode(obj.logo_droite))
                    f.close()
                    image_logo = Image.open(logo_path)  
                    width, height = image_logo.size 
                    max_height = entete_height
                    ratio = height/max_height
                    new_width = int(width/ratio)
                    logo_droite_resize = image_logo.resize((new_width, max_height))
                    logo_droite_width  = new_width

                #Ajouter le logo sauf pour le SVG qui est fait differement ************
                if file_extension!="svg":
                    #** Convertir la surface au format PIL pour ajouter le logo *******
                    def surface_to_pil(surface: cairo.ImageSurface) -> Image:
                        format = surface.get_format()
                        size = (surface.get_width(), surface.get_height())
                        stride = surface.get_stride()
                        with surface.get_data() as memory:
                            if format == cairo.Format.RGB24:
                                return Image.frombuffer(
                                    "RGB", size, memory.tobytes(),
                                    'raw', "BGRX", stride)
                            elif format == cairo.Format.ARGB32:
                                return Image.frombuffer(
                                    "RGBA", size, memory.tobytes(),
                                    'raw', "BGRa", stride)
                            else:
                                raise NotImplementedError(repr(format))

                    def pil_to_surface(image):
                        return cairo.ImageSurface.create_for_data(
                            bytearray(image.tobytes('raw', 'BGRa')),
                            cairo.FORMAT_ARGB32,
                            image.width,
                            image.height,
                            )

                    #** Combiner le logo et le gantt **********************************
                    result_pil = Image.new(
                        mode = 'RGBA',
                        size = (surface.get_width(), surface.get_height()),
                        color = (0, 0, 0, 0),
                        )
                    surface_pil = surface_to_pil(surface)    # Convertir le Gantt au format 'surface' au format PIL
                    result_pil.paste(surface_pil)            # Ajout du Gantt au format PIL
                    result_pil.paste(image_logo_resize)      # Ajout du logo au format PIL
                    if obj.logo_droite:
                        x = WIDTH - logo_droite_width
                        result_pil.paste(logo_droite_resize, (x, 0)) # Ajout du logo_droite au format PIL
                    surface = pil_to_surface(result_pil)     # Convertir le format PIL en format 'surface'


            if file_extension=="xlsx":
                obj.generer_xlsx_action(
                    path=path,items=items,date_debut=date_debut,date_fin=date_fin,jour_nb=jour_nb,
                    jour_fermeture_ids=jour_fermeture_ids,
                    titre=titre
                )


            #** Enregistrement du fichier ************************************* 
            if file_extension=="png":
                surface.write_to_png(path) # 2380 × 1684 pixels
                ctx.show_page()
                surface.finish()
                surface.flush()


            if file_extension=="pdf":
                # A4 portrait (point) : 595  , 842
                # A3 paysage (point)  : 595*2, 842
                # A2 portrait (point) : 595*2, 842*2
                # A1 paysage (point)  : 595*4, 842*2 (840 × 594 mm)
                pdf_surface = cairo.PDFSurface(path, WIDTH, HEIGHT) # 1 point = 0.3527777778 mm => https://www.unitconverters.net/length/point-to-millimeter.htm
                pdf_context = cairo.Context(pdf_surface)
                pdf_context.set_source_surface(surface)
                pdf_context.paint()
                pdf_context.show_page()
                pdf_surface.finish()
                pdf_surface.flush()

            if file_extension=="svg":
                surface.finish()
                surface.flush()

                #** PIL => base64 pour intégrer dans SVG **************************
                buff = BytesIO()
                image_logo_resize.save(buff, format="PNG")
                image_logo_base64 = base64.b64encode(buff.getvalue()).decode("utf-8")

                #** Ajout du logo dans le code SVG (XML) **********************
                #Source : https://docs.python.org/3/library/xml.etree.elementtree.html
                #Source : https://developer.mozilla.org/en-US/docs/Web/SVG/Element/circle
                ET.register_namespace("","http://www.w3.org/2000/svg")
                tree = ET.parse(path)
                root = tree.getroot()
                #ET.SubElement(root, "circle", cx="50", cy="50", r="100")
                href="data:image/png;base64,%s"%image_logo_base64
                ET.SubElement(root, "image", x="0", cy=str(entete_height), href=href)
                tree = ET.ElementTree(root)
                tree.write(path)

                #** Ajout du logo_droite dans le code SVG (XML) ***************
                if obj.logo_droite:
                    buff = BytesIO()
                    logo_droite_resize.save(buff, format="PNG")
                    image_logo_base64 = base64.b64encode(buff.getvalue()).decode("utf-8")
                    ET.register_namespace("","http://www.w3.org/2000/svg")
                    tree = ET.parse(path)
                    root = tree.getroot()
                    href="data:image/png;base64,%s"%image_logo_base64

                    x = WIDTH - logo_droite_width
                    ET.SubElement(root, "image", x=str(x), cy=str(entete_height), href=href)
                    tree = ET.ElementTree(root)
                    tree.write(path)

            # ** Creation ou modification de la pièce jointe ******************
            attachment_obj = self.env['ir.attachment']
            attachments = attachment_obj.search([('res_id','=',obj.id),('name','=',file_name_with_extension)])
            #if os.path.exists(path):
            #    raise ValidationError("Aucun fichier généré")
            if os.path.exists(path):
                datas = open(path,'rb').read()
                vals = {
                    'name':        file_name_with_extension,
                    'type':        'binary',
                    'res_model':   self._name,
                    'res_id':      obj.id,
                    'datas':       base64.b64encode(datas),
                }
                if attachments:
                    attachments.unlink()
                attachment = attachment_obj.with_context(image_no_postprocess=True).create(vals)
                os.unlink(path)
                msg = "Génération du %s"%obj.format_fichier.upper()
                obj.message_post(body=msg)
        
                #** Envoi du PDF dans le navigateur ************************************
                if attachment:
                    return {
                        'type' : 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true'%(attachment.id),
                    }
                #***********************************************************************


    def format_cell(self,cell,color=False,font_size=False):
        for obj in self:
            #cell = sheet.cell(row=row,column=i)
            font = copy(cell.font)
            if font_size:
                font.size = 12
            cell.font = font
            if color:
                fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            cell.fill=fill            


    def generer_xlsx_action(self,path=False,items=False,date_debut=False,date_fin=False,jour_nb=False,jour_fermeture_ids=False,titre=False):
        for obj in self:
            if not obj.modele_excel_id:
                raise Warning("Le modèle Excel n'est pas indiqué !")
            for modele in obj.modele_excel_id.modele_ids:
                #** Enregistrement du modèle en local *************************
                res = modele.datas
                res = base64.b64decode(res)
                f = open(path,'wb')
                f.write(res)
                f.close()
                #**************************************************************

                #** Chargement du classeur ************************************
                wb = load_workbook(path)
                sheet = wb.active

                #** Titre *****************************************************
                if titre:
                    sheet.cell(row=1, column=5).value = titre

                #** Semaines **************************************************
                semaines={}
                ladate = date_debut
                num_col_semaine=0
                for x in range(0,int(jour_nb)):
                    jour_semaine = ladate.isoweekday()
                    if x==0 or jour_semaine==1:
                        num_col_semaine+=1
                        semaine = ladate.strftime("%Y-S%V")
                        date_debut_semaine = ladate
                        nb_jours_semaine = 7 - jour_semaine
                        date_fin_semaine = ladate + timedelta(days=nb_jours_semaine)
                        if date_fin_semaine>date_fin:
                            date_fin_semaine=date_fin
                        semaines[semaine]={
                            'semaine'           : ladate.strftime("S%V"),
                            'date_debut_semaine': date_debut_semaine,
                            'date_fin_semaine'  : date_fin_semaine,
                            'num_col_semaine'   : num_col_semaine,
                        }
                    ladate += timedelta(days=1)
                i=0
                for key in semaines:
                    semaine = semaines[key]['semaine']
                    if semaine =='S01' or i==0:
                        date_fin_semaine = semaines[key]['date_fin_semaine']
                        sheet.cell(row=3, column=5+i).value = date_fin_semaine.strftime("%Y")
                    sheet.cell(row=4, column=5+i).value = semaine
                    i+=1
                nb_semaines = len(semaines)
                #**************************************************************

                #** Ajout des taches ******************************************
                row=5 #Première ligne pour enregistrer les données
                nb_taches=0
                for item in items:
                    niveau = item.get('niveau')
                    test=True
                    if date_fin and item['start_date']>date_fin:
                        test=False
                    if date_debut and item['end_date']<date_debut:
                        test=False
                    if niveau<2:
                        test=False
                    if test:
                        txt = '%s %s'%('- '*item.get('niveau'),item.get('text'))
                        sheet.cell(row=row, column=1).value = item.get('text')
                        sheet.cell(row=row, column=2).value = item.get('start_date')
                        sheet.cell(row=row, column=3).value = item.get('end_date')
                        avancement = "%s-%s"%(item['start_date'].strftime("S%V"),item['end_date'].strftime("S%V"))
                        sheet.cell(row=row, column=4).value = avancement

                        #Mettre un fond gris pour le niveau des sections ******
                        couleur_section = "e8e7e7" # GRis clair
                        if item.get('niveau')==2:
                            for i in  range(1,5):
                                cell = sheet.cell(row=row,column=i)
                                obj.format_cell(cell,color=couleur_section,font_size=12)

                        #** Couleur de la tache dans les semaines *************
                        border_color = (item.get('border_color') or '').lower()
                        color = _CSS_COLOR.get(border_color) or '#%s'%couleur_section
                        color = color.upper()[1:7]
                        i=0    


                        for key in semaines:
                            start_date = item['start_date']
                            end_date = item['end_date']
                            test=False
                            if start_date<=semaines[key]['date_fin_semaine'] and start_date>= semaines[key]['date_debut_semaine']:
                                test=True
                            if end_date<=semaines[key]['date_fin_semaine'] and end_date>= semaines[key]['date_debut_semaine']:
                                test=True
                            if start_date<=semaines[key]['date_debut_semaine'] and end_date>= semaines[key]['date_fin_semaine']:
                                test=True
                            if test:
                                cell = sheet.cell(row=row,column=i+5)
                                obj.format_cell(cell,color=color)
                            i+=1
                        #******************************************************
                        row+=1
                        nb_taches+=1

                #** Ajout des vacances ****************************************
                i=0
                if jour_fermeture_ids:
                    color = "c1c1c1" # Gris foncé
                    ladate = date_debut
                    for ct in range(0,jour_nb):
                        ladate_str = str(ladate)[0:10]
                        if ladate_str in jour_fermeture_ids:
                            semaine = ladate.strftime("%Y-S%V")
                            if semaine in semaines:
                                num_col_semaine = semaines[semaine]['num_col_semaine']
                                for lig in range(0,nb_taches):
                                    cell = sheet.cell(row=5+lig,column=num_col_semaine+4)
                                    obj.format_cell(cell,color=color)
                        ladate += timedelta(days=1)


                #** Zone d'impression *****************************************
                cell = sheet.cell(row=nb_taches+4,column=nb_semaines+4) # Cellule en bas à droite
                sheet.print_area = 'A1:%s'%cell.coordinate
                wb.save(path)
                #**************************************************************

